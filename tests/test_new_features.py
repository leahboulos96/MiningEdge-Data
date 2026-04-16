"""Tests for the new features added in this iteration:
  - remove a single scraper run / schedule run / bulk clear history
  - export filtered records as JSON / CSV / XLSX (UI + API)
  - live-activity endpoint & log capture during a scraper run
  - dynamic groups (create / edit / rename / delete) persisted in DB
"""

import io
import json


# ----------------- Helpers -----------------

def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _seed_record(db, status="pending", dedup="h1", source="news_afr",
                 title="Sample headline"):
    db.insert_record({
        "source": source, "source_group": "news", "record_type": "news",
        "external_id": "ext", "title": title, "description": "body",
        "dedup_hash": dedup,
    })
    rec = [r for r in db.list_records() if r["dedup_hash"] == dedup][0]
    if status != "pending":
        db.update_record_status(rec["id"], status)
    return rec


# ==================== Delete runs ====================

def test_delete_scraper_run(fresh_db):
    db = fresh_db
    rid = db.start_scraper_run("austender")
    db.finish_scraper_run(rid, "completed", items_found=1, items_new=1)
    assert len(db.recent_scraper_runs()) == 1
    db.delete_scraper_run(rid)
    assert len(db.recent_scraper_runs()) == 0


def test_delete_scraper_run_does_not_remove_running(fresh_db):
    """Actively running scrapers must NOT be deletable - that would orphan
    their live handler and break the UI."""
    db = fresh_db
    rid = db.start_scraper_run("austender")  # status stays 'running'
    db.delete_scraper_run(rid)
    assert len(db.recent_scraper_runs()) == 1


def test_bulk_clear_history_respects_scoping(fresh_db):
    db = fresh_db
    # Two finished runs for different scrapers
    for key in ["austender", "wa_tenders"]:
        rid = db.start_scraper_run(key)
        db.finish_scraper_run(rid, "completed")
    db.clear_scraper_runs(scraper="austender")
    remaining = [r["scraper"] for r in db.recent_scraper_runs()]
    assert remaining == ["wa_tenders"]


def test_delete_schedule_run_cascades(fresh_db):
    db = fresh_db
    sid = db.create_schedule("X", ["austender"], {"minute": "0", "hour": "0"})
    run_id = db.start_schedule_run(sid, "X")
    child = db.start_scraper_run("austender", schedule_id=sid,
                                 schedule_run_id=run_id)
    db.finish_scraper_run(child, "completed")
    db.finish_schedule_run(run_id, "ok", 1, 0, 0, [])
    assert len(db.recent_schedule_runs()) == 1
    db.delete_schedule_run(run_id)
    assert len(db.recent_schedule_runs()) == 0
    # Child scraper_run also gone
    assert all(r["id"] != child for r in db.recent_scraper_runs())


def test_delete_scraper_run_ui_route(app_client, fresh_db):
    client, _ = app_client
    rid = fresh_db.start_scraper_run("austender")
    fresh_db.finish_scraper_run(rid, "completed")
    r = client.post(f"/scraper-runs/{rid}/delete", follow_redirects=False)
    assert r.status_code in (302, 200)
    assert len(fresh_db.recent_scraper_runs()) == 0


def test_delete_schedule_run_ui_route(app_client, fresh_db):
    client, _ = app_client
    sid = fresh_db.create_schedule("s", ["austender"],
                                   {"minute": "0", "hour": "0"})
    rid = fresh_db.start_schedule_run(sid, "s")
    fresh_db.finish_schedule_run(rid, "ok", 0, 0, 0, [])
    client.post(f"/schedule-runs/{rid}/delete")
    assert len(fresh_db.recent_schedule_runs()) == 0


# ==================== Export records ====================

def test_export_json(fresh_db):
    import exports
    _seed_record(fresh_db, dedup="h-a", title="Alpha")
    _seed_record(fresh_db, dedup="h-b", title="Beta")
    body, mime, name = exports.build_export("json")
    data = json.loads(body)
    assert mime == "application/json" and name.endswith(".json")
    assert {r["title"] for r in data} == {"Alpha", "Beta"}


def test_export_csv_has_header_and_rows(fresh_db):
    import exports, csv
    _seed_record(fresh_db, dedup="h-a", title="Alpha")
    body, mime, _ = exports.build_export("csv")
    assert mime == "text/csv"
    rows = list(csv.DictReader(io.StringIO(body.decode("utf-8"))))
    assert len(rows) == 1 and rows[0]["title"] == "Alpha"


def test_export_xlsx_is_valid_workbook(fresh_db):
    import exports
    from openpyxl import load_workbook
    _seed_record(fresh_db, dedup="h-a", title="Alpha")
    body, mime, _ = exports.build_export("xlsx")
    assert "spreadsheet" in mime
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "title" in headers and "source" in headers
    assert ws.max_row == 2  # header + 1 data row


def test_export_filters_by_status(fresh_db):
    import exports
    _seed_record(fresh_db, dedup="h-a", title="Pend", status="pending")
    rec = _seed_record(fresh_db, dedup="h-b", title="App", status="pending")
    fresh_db.update_record_status(rec["id"], "approved")
    body, _, _ = exports.build_export("json", status="approved")
    data = json.loads(body)
    assert len(data) == 1 and data[0]["title"] == "App"


def test_export_bad_format_rejected(fresh_db):
    import exports, pytest
    with pytest.raises(ValueError):
        exports.build_export("pdf")


def test_ui_export_route(app_client, fresh_db):
    client, _ = app_client
    _seed_record(fresh_db, dedup="h-a")
    for fmt in ("json", "csv", "xlsx"):
        r = client.get(f"/records/export?format={fmt}&status=pending")
        assert r.status_code == 200
        assert r.headers["Content-Disposition"].startswith("attachment")


def test_api_export_requires_auth_and_returns_filtered(app_client, fresh_db):
    client, _ = app_client
    _seed_record(fresh_db, dedup="h-a", title="Alpha")
    rec = _seed_record(fresh_db, dedup="h-b", title="Beta")
    fresh_db.update_record_status(rec["id"], "approved")

    assert client.get("/api/v1/records/export?format=csv").status_code == 401

    token = fresh_db.create_api_key("exporter")
    r = client.get("/api/v1/records/export?format=json&status=approved",
                   headers=_auth(token))
    assert r.status_code == 200
    data = json.loads(r.data)
    assert len(data) == 1 and data[0]["title"] == "Beta"


# ==================== Live activity ====================

def test_live_endpoint_reports_running_and_captures_log(app_client, fresh_db,
                                                        monkeypatch):
    """Run a real scraper (a fake one) via the scheduler and assert that the
    /live.json endpoint returns its metadata + captured log lines."""
    import scheduler as sched_mod
    import registry
    from scrapers.base_scraper import BaseScraper

    class LogScraper(BaseScraper):
        name = "live_probe"
        source_group = "news"
        record_type = "news"
        def run(self):
            self.logger.info("probe starting")
            self.logger.info("probe midway")
            return []

    monkeypatch.setitem(registry.REGISTRY, "live_probe",
                        ("Live Probe", LogScraper, "news"))

    client, _ = app_client
    result = sched_mod.run_single_scraper("live_probe", triggered_by="unit")
    assert result["status"] == "ok"

    # While the run sits in the "recently finished" grace window, /live.json
    # should include it plus the log we emitted.
    r = client.get("/live.json")
    assert r.status_code == 200
    runs = r.get_json()["runs"]
    matches = [e for e in runs if e["scraper"] == "live_probe"]
    assert matches, "run not found in live feed"
    log = "\n".join(matches[0]["log"])
    assert "probe starting" in log
    assert "probe midway" in log


def test_api_live_requires_key(app_client, fresh_db):
    client, _ = app_client
    assert client.get("/api/v1/live").status_code == 401
    token = fresh_db.create_api_key("live")
    r = client.get("/api/v1/live", headers=_auth(token))
    assert r.status_code == 200
    assert "runs" in r.get_json()


# ==================== Dynamic groups ====================

def test_registry_defaults_used_when_no_custom(fresh_db):
    """With no DB-stored custom_groups, the default hardcoded groups apply."""
    import registry
    g = registry.groups()
    assert "tenders" in g and "news" in g and "asx" in g


def test_create_and_read_custom_group(fresh_db):
    import registry
    registry.save_group("my_news", ["news_afr", "news_west"])
    g = registry.groups()
    assert "my_news" in g
    assert g["my_news"] == ["news_afr", "news_west"]


def test_custom_group_filters_unknown_keys(fresh_db):
    import registry
    registry.save_group("bad", ["news_afr", "does_not_exist"])
    assert registry.groups()["bad"] == ["news_afr"]


def test_rename_group(fresh_db):
    import registry
    registry.save_group("old", ["news_afr"])
    registry.rename_group("old", "new")
    g = registry.groups()
    assert "new" in g and "old" not in g


def test_delete_group(fresh_db):
    import registry
    registry.save_group("tmp", ["news_afr"])
    registry.delete_group("tmp")
    assert "tmp" not in registry.groups()


def test_schedule_with_deleted_group_does_not_crash(fresh_db):
    """Schedules pointing at a now-deleted group resolve to [] - the
    scheduler must still run cleanly rather than blow up."""
    import registry, scheduler as sched_mod
    # Start with a custom group, schedule against it, delete it.
    registry.save_group("tempgrp", ["austender"])
    sid = fresh_db.create_schedule("s", ["group:tempgrp"],
                                   {"minute": "0", "hour": "0"})
    registry.delete_group("tempgrp")
    # Should not raise - resolution returns empty list, the run ends ok.
    sched_mod.run_schedule(sid)
    runs = fresh_db.recent_schedule_runs(limit=1)
    assert runs and runs[0]["status"] in ("ok", "error")


def test_groups_page_renders(app_client, fresh_db):
    client, _ = app_client
    assert client.get("/groups").status_code == 200
    assert client.get("/groups/new").status_code == 200


def test_wipe_all_records_requires_confirm(app_client, fresh_db):
    client, _ = app_client
    _seed_record(fresh_db, dedup="h1")
    # No confirm field -> should NOT wipe
    client.post("/records/wipe-all", data={})
    assert fresh_db.count_records() == 1
    # With confirm -> wipes
    r = client.post("/records/wipe-all", data={"confirm": "yes"},
                    follow_redirects=False)
    assert r.status_code in (302, 200)
    assert fresh_db.count_records() == 0


def test_wipe_all_records_can_clear_discarded_keys(app_client, fresh_db):
    client, _ = app_client
    rec = _seed_record(fresh_db, dedup="h-kill")
    fresh_db.update_record_status(rec["id"], "discarded")
    assert fresh_db.is_discarded("h-kill")

    # Default wipe: records gone, discarded keys kept
    client.post("/records/wipe-all", data={"confirm": "yes"})
    assert fresh_db.is_discarded("h-kill")

    # Re-insert, re-discard, then wipe with checkbox -> keys also cleared
    _seed_record(fresh_db, dedup="h-kill-2")
    rec2 = fresh_db.list_records()[0]
    fresh_db.update_record_status(rec2["id"], "discarded")
    client.post("/records/wipe-all",
                data={"confirm": "yes", "clear_discarded_keys": "on"})
    assert not fresh_db.is_discarded("h-kill")
    assert not fresh_db.is_discarded("h-kill-2")


def test_wipe_all_scraper_runs(app_client, fresh_db):
    client, _ = app_client
    rid = fresh_db.start_scraper_run("austender")
    fresh_db.finish_scraper_run(rid, "completed")
    client.post("/scraper-runs/wipe-all", data={"confirm": "yes"})
    assert len(fresh_db.recent_scraper_runs()) == 0


def test_wipe_all_schedule_runs_cascades(app_client, fresh_db):
    client, _ = app_client
    sid = fresh_db.create_schedule("s", ["austender"],
                                   {"minute": "0", "hour": "0"})
    rid = fresh_db.start_schedule_run(sid, "s")
    child = fresh_db.start_scraper_run("austender", schedule_id=sid,
                                       schedule_run_id=rid)
    fresh_db.finish_scraper_run(child, "completed")
    fresh_db.finish_schedule_run(rid, "ok", 1, 0, 0, [])
    client.post("/schedule-runs/wipe-all", data={"confirm": "yes"})
    assert len(fresh_db.recent_schedule_runs()) == 0
    assert all(r["schedule_run_id"] != rid
               for r in fresh_db.recent_scraper_runs())


def test_wipe_all_logs(app_client, monkeypatch, tmp_path):
    client, app_mod = app_client
    import config, os
    # write fake log files
    for name in ("a.log", "b.log"):
        with open(os.path.join(config.LOGS_DIR, name), "w") as f:
            f.write("hi")
    # without confirm - not deleted
    client.post("/logs/wipe-all", data={})
    assert any(f.endswith(".log") for f in os.listdir(config.LOGS_DIR))
    # with confirm - cleared
    client.post("/logs/wipe-all", data={"confirm": "yes"})
    assert not any(f.endswith(".log") for f in os.listdir(config.LOGS_DIR))


def test_records_page_shows_api_url(app_client, fresh_db):
    client, _ = app_client
    r = client.get("/records?status=approved&source=austender")
    body = r.data.decode("utf-8")
    assert "/api/v1/records?status=approved" in body
    assert "/api/v1/records/export?format=csv&amp;status=approved" in body
    assert 'id="api-list-url"' in body
    assert 'id="api-export-url"' in body


def test_group_create_edit_delete_ui(app_client, fresh_db):
    client, _ = app_client
    import registry

    # Create
    client.post("/groups/new", data={"name": "vic_news",
                                     "scrapers": ["news_afr"]})
    assert "vic_news" in registry.groups()

    # Edit (add a scraper + rename)
    client.post("/groups/vic_news/edit",
                data={"name": "oz_news",
                      "scrapers": ["news_afr", "news_west"]})
    g = registry.groups()
    assert "oz_news" in g and "vic_news" not in g
    assert set(g["oz_news"]) == {"news_afr", "news_west"}

    # Delete
    client.post("/groups/oz_news/delete")
    assert "oz_news" not in registry.groups()

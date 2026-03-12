"""
Flask Dashboard for managing Australian Tender & ASX scrapers.
Features: Login, dashboard, manual run, logs viewer, settings, scheduling.
"""

import os
import io
import re
import csv
import json
import zipfile
import threading
from datetime import datetime
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, Response
from apscheduler.schedulers.background import BackgroundScheduler

import config
from scrapers.tenders.austender import AusTenderScraper
from scrapers.tenders.wa_tenders import WATendersScraper
from scrapers.tenders.qld_tenders import QLDTendersScraper
from scrapers.tenders.sa_tenders import SATendersScraper
from scrapers.tenders.icn_gateway import ICNGatewayScraper
from scrapers.tenders.icn_workpackages import ICNWorkpackagesScraper
from scrapers.asx.asx_scraper import ASXScraper

app = Flask(__name__)
app.secret_key = config.FLASK_SECRET_KEY

# State tracking
scraper_status = {}
running_scrapers = {}
scheduler = BackgroundScheduler()
scheduler.start()

SCRAPER_MAP = {
    "austender": ("AusTender (Federal)", AusTenderScraper),
    "wa_tenders": ("WA Tenders", WATendersScraper),
    "qld_tenders": ("QLD QTenders", QLDTendersScraper),
    "sa_tenders": ("SA Tenders", SATendersScraper),
    "icn_gateway": ("ICN Gateway (Projects)", ICNGatewayScraper),
    "icn_workpackages": ("ICN Gateway (Work Packages)", ICNWorkpackagesScraper),
    "asx_announcements": ("ASX Announcements", ASXScraper),
}

# Settings file path
SETTINGS_FILE = os.path.join(config.BASE_DIR, "settings.json")


def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "r") as f:
            return json.load(f)
    return {
        "scrape_do_token": config.SCRAPE_DO_TOKEN,
        "asx_tickers": config.ASX_TICKERS,
        "schedule_enabled": False,
        "schedule_hour": 6,
        "schedule_minute": 0,
        "enabled_scrapers": list(SCRAPER_MAP.keys()),
    }


def save_settings(settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings, f, indent=2)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def get_output_files():
    files = []
    if os.path.exists(config.OUTPUT_DIR):
        for fn in sorted(os.listdir(config.OUTPUT_DIR), reverse=True):
            if fn.endswith(".json"):
                path = os.path.join(config.OUTPUT_DIR, fn)
                size = os.path.getsize(path)
                mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
                try:
                    with open(path, "r") as f:
                        data = json.load(f)
                        count = len(data) if isinstance(data, list) else 0
                except Exception:
                    count = 0
                files.append({"name": fn, "size": size, "modified": mtime, "records": count})
    return files


def get_log_files():
    files = []
    if os.path.exists(config.LOGS_DIR):
        for fn in sorted(os.listdir(config.LOGS_DIR), reverse=True):
            if fn.endswith(".log"):
                path = os.path.join(config.LOGS_DIR, fn)
                size = os.path.getsize(path)
                mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
                files.append({"name": fn, "size": size, "modified": mtime})
    return files


def run_scraper_background(scraper_key):
    """Run a scraper in a background thread."""
    if scraper_key in running_scrapers and running_scrapers[scraper_key]:
        return

    label, scraper_cls = SCRAPER_MAP[scraper_key]
    running_scrapers[scraper_key] = True
    scraper_status[scraper_key] = {
        "status": "running",
        "started": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "items": 0,
        "errors": 0,
    }

    def _run():
        try:
            # Apply current settings
            settings = load_settings()
            config.SCRAPE_DO_TOKEN = settings.get("scrape_do_token", config.SCRAPE_DO_TOKEN)
            config.ASX_TICKERS = settings.get("asx_tickers", config.ASX_TICKERS)

            scraper = scraper_cls()
            results = scraper.execute()
            scraper_status[scraper_key] = {
                "status": "completed",
                "started": scraper_status[scraper_key]["started"],
                "finished": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "items": len(results) if results else 0,
                "errors": scraper.stats.get("errors", 0),
            }
        except Exception as e:
            scraper_status[scraper_key] = {
                "status": "error",
                "started": scraper_status[scraper_key]["started"],
                "finished": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "items": 0,
                "errors": 1,
                "error_message": str(e),
            }
        finally:
            running_scrapers[scraper_key] = False

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()


def run_all_background():
    """Run all enabled scrapers sequentially in background."""
    settings = load_settings()
    enabled = settings.get("enabled_scrapers", list(SCRAPER_MAP.keys()))
    for key in enabled:
        if key in SCRAPER_MAP:
            run_scraper_background(key)


# --- Routes ---

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == config.ADMIN_USERNAME and password == config.ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("dashboard"))
        flash("Invalid credentials", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    output_files = get_output_files()
    return render_template(
        "dashboard.html",
        scrapers=SCRAPER_MAP,
        scraper_status=scraper_status,
        running_scrapers=running_scrapers,
        output_files=output_files,
    )


@app.route("/run/<scraper_key>")
@login_required
def run_scraper(scraper_key):
    if scraper_key == "all":
        run_all_background()
        flash("All scrapers started", "success")
    elif scraper_key in SCRAPER_MAP:
        if running_scrapers.get(scraper_key):
            flash(f"{SCRAPER_MAP[scraper_key][0]} is already running", "warning")
        else:
            run_scraper_background(scraper_key)
            flash(f"{SCRAPER_MAP[scraper_key][0]} started", "success")
    else:
        flash("Unknown scraper", "error")
    return redirect(url_for("dashboard"))


@app.route("/status")
@login_required
def status_api():
    return jsonify({
        "scraper_status": scraper_status,
        "running": {k: v for k, v in running_scrapers.items() if v},
    })


@app.route("/logs")
@login_required
def logs():
    log_files = get_log_files()
    selected = request.args.get("file", "")
    content = ""
    if selected:
        path = os.path.join(config.LOGS_DIR, selected)
        if os.path.exists(path) and selected.endswith(".log"):
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
    return render_template("logs.html", log_files=log_files, selected=selected, content=content)


@app.route("/download/<filename>")
@login_required
def download_output(filename):
    path = os.path.join(config.OUTPUT_DIR, filename)
    if os.path.exists(path) and filename.endswith(".json"):
        return send_file(path, as_attachment=True)
    flash("File not found", "error")
    return redirect(url_for("dashboard"))


@app.route("/download-csv/<filename>")
@login_required
def download_csv(filename):
    path = os.path.join(config.OUTPUT_DIR, filename)
    if not os.path.exists(path) or not filename.endswith(".json"):
        flash("File not found", "error")
        return redirect(url_for("dashboard"))

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list) or not data:
        flash("No data to export", "error")
        return redirect(url_for("dashboard"))

    # Collect all keys across all records
    all_keys = []
    seen = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        writer.writerow({k: str(v) if v is not None else "" for k, v in row.items()})

    csv_name = filename.replace(".json", ".csv")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={csv_name}"},
    )


@app.route("/download-xlsx/<filename>")
@login_required
def download_xlsx(filename):
    path = os.path.join(config.OUTPUT_DIR, filename)
    if not os.path.exists(path) or not filename.endswith(".json"):
        flash("File not found", "error")
        return redirect(url_for("dashboard"))

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list) or not data:
        flash("No data to export", "error")
        return redirect(url_for("dashboard"))

    from openpyxl import Workbook

    # Collect all keys across all records
    all_keys = []
    seen = set()
    for row in data:
        for k in row.keys():
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row
    for col, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col, key in enumerate(all_keys, 1):
            val = row.get(key, "")
            ws.cell(row=row_idx, column=col, value=str(val) if val is not None else "")

    # Auto-width columns
    for col, key in enumerate(all_keys, 1):
        max_len = len(key)
        for row_idx in range(2, min(len(data) + 2, 52)):  # sample first 50 rows
            cell_val = str(ws.cell(row=row_idx, column=col).value or "")
            max_len = max(max_len, min(len(cell_val), 50))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    xlsx_name = filename.replace(".json", ".xlsx")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=xlsx_name,
    )


@app.route("/view/<filename>")
@login_required
def view_output(filename):
    path = os.path.join(config.OUTPUT_DIR, filename)
    if os.path.exists(path) and filename.endswith(".json"):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        pretty = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        count = len(data) if isinstance(data, list) else 0
        return render_template("view_json.html", filename=filename, content=pretty, count=count)
    flash("File not found", "error")
    return redirect(url_for("dashboard"))


@app.route("/delete/output/<filename>")
@login_required
def delete_output(filename):
    path = os.path.join(config.OUTPUT_DIR, filename)
    if os.path.exists(path) and filename.endswith(".json"):
        os.remove(path)
        flash(f"Deleted {filename}", "success")
    else:
        flash("File not found", "error")
    return redirect(url_for("dashboard"))


@app.route("/delete/output-all")
@login_required
def delete_all_output():
    count = 0
    for fn in os.listdir(config.OUTPUT_DIR):
        if fn.endswith(".json"):
            os.remove(os.path.join(config.OUTPUT_DIR, fn))
            count += 1
    flash(f"Deleted {count} output files", "success")
    return redirect(url_for("dashboard"))


@app.route("/delete/log/<filename>")
@login_required
def delete_log(filename):
    path = os.path.join(config.LOGS_DIR, filename)
    if os.path.exists(path) and filename.endswith(".log"):
        os.remove(path)
        flash(f"Deleted {filename}", "success")
    else:
        flash("File not found", "error")
    return redirect(url_for("logs"))


@app.route("/delete/logs-all")
@login_required
def delete_all_logs():
    count = 0
    for fn in os.listdir(config.LOGS_DIR):
        if fn.endswith(".log"):
            os.remove(os.path.join(config.LOGS_DIR, fn))
            count += 1
    flash(f"Deleted {count} log files", "success")
    return redirect(url_for("logs"))


def load_icn_cookies():
    """Load ICN Gateway cookies from icn_cookies.json."""
    cookie_path = config.ICN_COOKIES_FILE
    if os.path.exists(cookie_path):
        try:
            with open(cookie_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {}


def save_icn_cookies(cookies):
    """Save ICN Gateway cookies to icn_cookies.json."""
    cookie_path = config.ICN_COOKIES_FILE
    try:
        with open(cookie_path, "w", encoding="utf-8") as f:
            json.dump(cookies, f, indent=2)
    except IOError:
        pass


@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    current = load_settings()

    if request.method == "POST":
        current["scrape_do_token"] = request.form.get("scrape_do_token", current["scrape_do_token"])

        tickers_raw = request.form.get("asx_tickers", "")
        if tickers_raw.strip():
            current["asx_tickers"] = [t.strip().upper() for t in tickers_raw.replace("\n", ",").split(",") if t.strip()]

        # Save ICN Gateway cookies
        icn_phpsessid = request.form.get("icn_phpsessid", "").strip()
        icn_remember_tfa = request.form.get("icn_remember_tfa", "").strip()
        icn_xsrf = request.form.get("icn_xsrf", "").strip()
        icn_session = request.form.get("icn_session", "").strip()

        # Only write cookies file if at least PHPSESSID and session are provided
        if icn_phpsessid and icn_session:
            icn_cookies = {
                "PHPSESSID": icn_phpsessid,
                "remember_tfa_gateway": icn_remember_tfa,
                "XSRF-TOKEN": icn_xsrf,
                "gateway_by_icn_session": icn_session,
            }
            # Remove empty values
            icn_cookies = {k: v for k, v in icn_cookies.items() if v}
            save_icn_cookies(icn_cookies)

        current["schedule_enabled"] = "schedule_enabled" in request.form
        current["schedule_hour"] = int(request.form.get("schedule_hour", 6))
        current["schedule_minute"] = int(request.form.get("schedule_minute", 0))

        enabled = request.form.getlist("enabled_scrapers")
        if enabled:
            current["enabled_scrapers"] = enabled

        save_settings(current)

        # Apply settings
        config.SCRAPE_DO_TOKEN = current["scrape_do_token"]
        config.ASX_TICKERS = current["asx_tickers"]

        # Update scheduler
        _update_schedule(current)

        flash("Settings saved", "success")
        return redirect(url_for("settings"))

    # Load ICN cookies for display
    icn_cookies = load_icn_cookies()

    return render_template("settings.html", settings=current, scrapers=SCRAPER_MAP, icn_cookies=icn_cookies)


# --- Backup / Restore ---

def _get_available_months():
    """Scan output and log files to find available YYYY-MM months."""
    months = set()
    date_pattern = re.compile(r"(\d{4})[-_]?(\d{2})[-_]?\d{2}")
    for d in (config.OUTPUT_DIR, config.LOGS_DIR):
        if os.path.exists(d):
            for fn in os.listdir(d):
                m = date_pattern.search(fn)
                if m:
                    months.add(f"{m.group(1)}-{m.group(2)}")
    return sorted(months, reverse=True)


def _get_available_scrapers():
    """Scan output files to find scraper prefixes that have data."""
    scrapers = set()
    if os.path.exists(config.OUTPUT_DIR):
        for fn in os.listdir(config.OUTPUT_DIR):
            if fn.endswith(".json"):
                # filename pattern: scrapername_YYYYMMDD.json
                parts = fn.rsplit("_", 1)
                if len(parts) == 2:
                    scrapers.add(parts[0])
    return sorted(scrapers)


def _file_matches_filters(filename, month_filter, scraper_filter):
    """Check if a filename matches the selected month and/or scraper filters."""
    if month_filter:
        date_pattern = re.compile(r"(\d{4})[-_]?(\d{2})[-_]?\d{2}")
        m = date_pattern.search(filename)
        if not m or f"{m.group(1)}-{m.group(2)}" != month_filter:
            return False
    if scraper_filter:
        parts = filename.rsplit("_", 1)
        if len(parts) < 2 or parts[0] != scraper_filter:
            return False
    return True


@app.route("/backup")
@login_required
def backup():
    months = _get_available_months()
    scrapers = _get_available_scrapers()
    output_count = len([f for f in os.listdir(config.OUTPUT_DIR) if f.endswith(".json")]) if os.path.exists(config.OUTPUT_DIR) else 0
    log_count = len([f for f in os.listdir(config.LOGS_DIR) if f.endswith(".log")]) if os.path.exists(config.LOGS_DIR) else 0
    has_settings = os.path.exists(SETTINGS_FILE)
    has_cookies = os.path.exists(config.ICN_COOKIES_FILE)
    return render_template(
        "backup.html",
        months=months,
        scrapers=scrapers,
        output_count=output_count,
        log_count=log_count,
        has_settings=has_settings,
        has_cookies=has_cookies,
    )


@app.route("/backup/export", methods=["POST"])
@login_required
def backup_export():
    month_filter = request.form.get("month", "").strip()
    scraper_filter = request.form.get("scraper", "").strip()
    include_settings = "include_settings" in request.form
    include_cookies = "include_cookies" in request.form
    include_logs = "include_logs" in request.form

    buf = io.BytesIO()
    file_count = 0

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Output files
        if os.path.exists(config.OUTPUT_DIR):
            for fn in sorted(os.listdir(config.OUTPUT_DIR)):
                if fn.endswith(".json") and _file_matches_filters(fn, month_filter, scraper_filter):
                    zf.write(os.path.join(config.OUTPUT_DIR, fn), f"output/{fn}")
                    file_count += 1

        # Log files
        if include_logs and os.path.exists(config.LOGS_DIR):
            for fn in sorted(os.listdir(config.LOGS_DIR)):
                if fn.endswith(".log") and _file_matches_filters(fn, month_filter, scraper_filter):
                    zf.write(os.path.join(config.LOGS_DIR, fn), f"logs/{fn}")
                    file_count += 1

        # Settings
        if include_settings and os.path.exists(SETTINGS_FILE):
            zf.write(SETTINGS_FILE, "settings.json")
            file_count += 1

        # ICN cookies
        if include_cookies and os.path.exists(config.ICN_COOKIES_FILE):
            zf.write(config.ICN_COOKIES_FILE, "icn_cookies.json")
            file_count += 1

    if file_count == 0:
        flash("No files matched the selected filters", "warning")
        return redirect(url_for("backup"))

    buf.seek(0)

    # Build descriptive filename
    parts = ["miningedge_backup"]
    if scraper_filter:
        parts.append(scraper_filter)
    if month_filter:
        parts.append(month_filter)
    parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    zip_name = "_".join(parts) + ".zip"

    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name=zip_name)


@app.route("/backup/import", methods=["POST"])
@login_required
def backup_import():
    uploaded = request.files.get("backup_file")
    if not uploaded or not uploaded.filename.endswith(".zip"):
        flash("Please upload a .zip file", "error")
        return redirect(url_for("backup"))

    imported = {"output": 0, "logs": 0, "settings": False, "cookies": False}

    try:
        with zipfile.ZipFile(uploaded.stream, "r") as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue

                name = info.filename
                # Sanitize: skip anything with path traversal
                if ".." in name or name.startswith("/"):
                    continue

                if name.startswith("output/") and name.endswith(".json"):
                    fn = os.path.basename(name)
                    target = os.path.join(config.OUTPUT_DIR, fn)
                    with zf.open(info) as src, open(target, "wb") as dst:
                        dst.write(src.read())
                    imported["output"] += 1

                elif name.startswith("logs/") and name.endswith(".log"):
                    fn = os.path.basename(name)
                    target = os.path.join(config.LOGS_DIR, fn)
                    with zf.open(info) as src, open(target, "wb") as dst:
                        dst.write(src.read())
                    imported["logs"] += 1

                elif name == "settings.json":
                    with zf.open(info) as src:
                        data = json.loads(src.read().decode("utf-8"))
                    save_settings(data)
                    config.SCRAPE_DO_TOKEN = data.get("scrape_do_token", config.SCRAPE_DO_TOKEN)
                    config.ASX_TICKERS = data.get("asx_tickers", config.ASX_TICKERS)
                    _update_schedule(data)
                    imported["settings"] = True

                elif name == "icn_cookies.json":
                    with zf.open(info) as src:
                        data = json.loads(src.read().decode("utf-8"))
                    save_icn_cookies(data)
                    imported["cookies"] = True

        parts = []
        if imported["output"]:
            parts.append(f"{imported['output']} output files")
        if imported["logs"]:
            parts.append(f"{imported['logs']} log files")
        if imported["settings"]:
            parts.append("settings")
        if imported["cookies"]:
            parts.append("ICN cookies")
        flash(f"Imported: {', '.join(parts)}", "success")

    except zipfile.BadZipFile:
        flash("Invalid zip file", "error")
    except Exception as e:
        flash(f"Import error: {str(e)}", "error")

    return redirect(url_for("backup"))


def _update_schedule(settings):
    """Update the APScheduler job based on settings."""
    job_id = "daily_scrape"
    try:
        scheduler.remove_job(job_id)
    except Exception:
        pass

    if settings.get("schedule_enabled"):
        scheduler.add_job(
            run_all_background,
            "cron",
            hour=settings.get("schedule_hour", 6),
            minute=settings.get("schedule_minute", 0),
            id=job_id,
        )


# Apply saved settings on startup
_saved = load_settings()
config.SCRAPE_DO_TOKEN = _saved.get("scrape_do_token", config.SCRAPE_DO_TOKEN)
config.ASX_TICKERS = _saved.get("asx_tickers", config.ASX_TICKERS)
_update_schedule(_saved)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)

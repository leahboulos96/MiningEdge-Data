"""
Export filtered records as JSON / CSV / XLSX. Used by both the web UI
(/records/export) and the REST API (/api/v1/records/export).
"""

import io
import csv
import json
from datetime import datetime

import db


EXPORT_COLUMNS = [
    "id", "source", "source_group", "record_type", "external_id",
    "title", "description", "entity_name", "region",
    "published_date", "closing_date", "url", "pdf_url",
    "status", "scraped_at", "reviewed_at", "reviewed_by", "review_notes",
]


def fetch_records(status=None, source=None, search=None):
    """Return the full filtered list - no pagination (export = everything)."""
    # Large hard-cap so a runaway export doesn't eat memory on an unbounded DB.
    return db.list_records(status=status, source=source, search=search,
                           limit=100_000, offset=0)


def _slice(records):
    return [{k: r.get(k) for k in EXPORT_COLUMNS} for r in records]


def to_json(records):
    return json.dumps(_slice(records), indent=2, ensure_ascii=False,
                      default=str).encode("utf-8")


def to_csv(records):
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    w.writeheader()
    for r in _slice(records):
        w.writerow({k: ("" if v is None else str(v)) for k, v in r.items()})
    return out.getvalue().encode("utf-8")


def to_xlsx(records):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "records"
    for col, key in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = cell.font.copy(bold=True)
    for row_idx, r in enumerate(_slice(records), 2):
        for col, key in enumerate(EXPORT_COLUMNS, 1):
            v = r.get(key)
            ws.cell(row=row_idx, column=col,
                    value="" if v is None else str(v))
    for col, key in enumerate(EXPORT_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = \
            min(max(len(key) + 2, 14), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_export(fmt, status=None, source=None, search=None):
    """Return (bytes, mimetype, filename). Raises ValueError for bad format."""
    fmt = (fmt or "json").lower()
    records = fetch_records(status=status, source=source, search=search)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    parts = ["records", status or "all"]
    if source:
        parts.append(source)
    base = "_".join(parts) + f"_{stamp}"

    if fmt == "json":
        return to_json(records), "application/json", f"{base}.json"
    if fmt == "csv":
        return to_csv(records), "text/csv", f"{base}.csv"
    if fmt == "xlsx":
        return (
            to_xlsx(records),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{base}.xlsx",
        )
    raise ValueError(f"unknown format: {fmt}")
